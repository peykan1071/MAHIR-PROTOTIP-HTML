"""Read teacher-made score tables directly from modern Excel workbooks."""

from __future__ import annotations

import io

from .table_parser import parse_tabular_document


def parse_score_xlsx(content: bytes) -> dict[str, object]:
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise ValueError("Excel okuma bileşeni kurulu değil. 'pip install openpyxl' komutunu çalıştırınız.") from error

    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        tables = [list(worksheet.iter_rows(values_only=True)) for worksheet in workbook.worksheets]
        workbook.close()
    except Exception as error:
        raise ValueError("Excel belgesi açılamadı veya geçerli bir .xlsx dosyası değil.") from error

    return parse_tabular_document(tables, source_label="Excel belgesi")
